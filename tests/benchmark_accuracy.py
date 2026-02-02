#!/usr/bin/env python3
"""
PII Detection Accuracy Benchmark

Modes:
- quick: Training_Set.xlsx only (500 samples)
- full:  Training_Set.xlsx + Generated test files + CSV vs PDF comparison

Usage:
    python benchmark_accuracy.py          # quick test
    python benchmark_accuracy.py --full   # full test
    python benchmark_accuracy.py --samples 1000  # custom sample size
"""

import re
import csv
import random
import time
import sys
import argparse
import subprocess
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

# ============================================================================
# PII DETECTION PATTERNS (Improved)
# ============================================================================
PATTERNS = {
    # EMAIL: Improved to handle more TLDs and subdomains
    'EMAIL': re.compile(
        r'\b[A-Za-z0-9._%+-]+@'
        r'[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?'
        r'(?:\.[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?)*'
        r'\.[A-Za-z]{2,}\b'
    ),
    # PHONE: International formats
    'PHONE': re.compile(r'\b(?:\+\d{1,3}[-.\s]?)?\(?\d{1,4}\)?[-.\s]?\d{3,5}[-.\s]?\d{4}\b'),
    # SSN
    'SSN': re.compile(r'\b\d{3}[-\s]?\d{2}[-\s]?\d{4}\b'),
    # URL
    'URL': re.compile(r'https?://[^\s<>"{}|\\^`\[\]]+|www\.[^\s<>"{}|\\^`\[\]]+'),
    # CREDIT_CARD: Added JCB (35xx), Diners (30x, 36, 38), UnionPay (62)
    'CREDIT_CARD': re.compile(
        r'\b(?:'
        r'4\d{3}|'                    # Visa
        r'5[1-5]\d{2}|'               # Mastercard
        r'6011|65\d{2}|64[4-9]\d|'    # Discover
        r'3[47]\d{2}|'                # Amex
        r'35(?:2[89]|[3-8]\d)|'       # JCB (3528-3589)
        r'30[0-5]\d|36\d{2}|38\d{2}|' # Diners Club
        r'62\d{2}'                    # UnionPay
        r')[-\s]?\d{4}[-.\s]?\d{4}[-.\s]?\d{2,4}\b'
    ),
    # PERSON: Context-aware (title or label required)
    'PERSON': re.compile(
        r'\b(?:Dr|Mr|Mrs|Ms|Prof|Rev|Sir|Lady|Lord|Capt|Col|Gen|Sgt|Rabbi|Father)\.?\s+'
        r'[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?\b|'
        r'[Nn]ame[:\s]+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)'
    ),
    # ADDRESS: More street suffixes and Suite/Apt handling
    'ADDRESS': re.compile(
        r'\b\d{1,5}\s+[A-Z][a-z]+(?:\s+[A-Za-z]+)*\s+'
        r'(?:Street|St|Avenue|Ave|Road|Rd|Drive|Dr|Lane|Ln|Boulevard|Blvd|'
        r'Way|Court|Ct|Place|Pl|Circle|Cir|Terrace|Ter|Trail|Trl|'
        r'Parkway|Pkwy|Highway|Hwy|Falls|River|Creek|Ridge|Hill|Valley|'
        r'Park|Grove|Square|Plaza|Point|View|Heights|Meadow|Garden)'
        r'(?:\s+(?:Suite|Ste|Apt|Unit|#)\s*\w+)?',
        re.IGNORECASE
    ),
    # COMPANY: Fixed to avoid catastrophic backtracking - limit word count and length
    'COMPANY': re.compile(
        r'\b[A-Z][A-Za-z]{1,30}(?:[-\s&][A-Z]?[A-Za-z]{1,30}){0,5}\s+'
        r'(?:Ltd|LLC|Inc|Corp|Co|Company|GmbH|PLC|LP|LLP|'
        r'Partners|Group|Holdings|International|Enterprises|'
        r'Associates|Services|Solutions|Technologies|Systems)\b',
        re.IGNORECASE
    ),
    # CURRENCY
    'CURRENCY': re.compile(r'[\$\xa3\u20ac\xa5\u20b9]\s*[\d,]+(?:\.\d{2})?'),
}

# Field mapping: Pattern name -> Excel column name
FIELD_MAP = {
    'EMAIL': 'Email',
    'PHONE': 'Phone',
    'SSN': 'SSN',
    'URL': 'URL',
    'CREDIT_CARD': 'Credit Card',
    'PERSON': 'Name',
    'ADDRESS': 'Address',
    'COMPANY': 'Company',
}


def detect_pii(text):
    """Detect PII in text using regex patterns."""
    if not text:
        return {}
    results = {}
    for pii_type, pattern in PATTERNS.items():
        matches = pattern.findall(str(text))
        flat = []
        for m in matches:
            if isinstance(m, tuple):
                flat.extend([x for x in m if x])
            else:
                flat.append(m)
        if flat:
            results[pii_type] = flat
    return results


def normalize(text):
    """Normalize text for comparison."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', str(text).lower().strip())


def check_match(detected_list, ground_truth):
    """Check if ground truth was detected."""
    if not ground_truth:
        return None
    gt = normalize(ground_truth)
    if not gt or len(gt) < 3:
        return None
    for det in detected_list:
        d = normalize(det)
        if gt in d or d in gt:
            return True
        # Partial word match
        gt_words = set(gt.split())
        d_words = set(d.split())
        if gt_words and len(gt_words & d_words) >= len(gt_words) * 0.5:
            return True
    return False


def run_training_set_benchmark(sample_size=500):
    """Run benchmark against Training_Set.xlsx ground truth."""
    data_dir = Path(__file__).parent / 'data' / 'training'
    xlsx_path = data_dir / 'Training_Set.xlsx'
    cache_path = data_dir / 'Training_Set_cache.csv'

    # Try to use cached CSV first (much faster)
    if cache_path.exists():
        print(f"\nLoading from cache: {cache_path.name}...")
        rows = []
        with open(cache_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            max_load = min(sample_size * 2, 5000)
            for i, row in enumerate(reader):
                if i >= max_load:
                    break
                rows.append(row)
    elif xlsx_path.exists():
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl required for Training_Set benchmark")
            return None

        print(f"\nLoading Training_Set.xlsx (this may take ~15s on first run)...")
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        sheet = wb.active

        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        # Load rows (limit to reasonable size)
        rows = []
        max_load = min(sample_size * 2, 5000)
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if i >= max_load:
                break
            rows.append(dict(zip(headers, row)))

        wb.close()

        # Create cache for faster subsequent runs
        print(f"Creating cache file for faster future runs...")
        with open(cache_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
    else:
        print(f"Error: {xlsx_path} not found")
        return None

    # Sample
    sample = random.sample(rows, min(sample_size, len(rows)))
    print(f"Sampled: {len(sample)} rows from {len(rows)} loaded\n")

    # Track accuracy
    accuracy = {k: {'tp': 0, 'fn': 0, 'total': 0} for k in FIELD_MAP.keys()}

    start = time.time()

    for row in sample:
        text = row.get('Text', '')
        detections = detect_pii(text)

        for pii_type, field in FIELD_MAP.items():
            gt_value = row.get(field, '')
            detected_list = detections.get(pii_type, [])

            if gt_value and str(gt_value).strip():
                accuracy[pii_type]['total'] += 1
                result = check_match(detected_list, gt_value)
                if result:
                    accuracy[pii_type]['tp'] += 1
                else:
                    accuracy[pii_type]['fn'] += 1

    elapsed = time.time() - start

    return {
        'source': 'Training_Set.xlsx',
        'samples': len(sample),
        'accuracy': accuracy,
        'time': elapsed,
    }


def run_generated_benchmark():
    """Run benchmark on generated test files (detection count only)."""
    generated_dir = Path(__file__).parent / 'data' / 'generated'
    if not generated_dir.exists():
        print(f"Warning: {generated_dir} not found")
        return None

    csv_files = sorted(generated_dir.glob('Test_*.csv'))
    if not csv_files:
        print("No generated test files found")
        return None

    print(f"\nProcessing {len(csv_files)} generated test files...")

    total_stats = {
        'texts': 0,
        'detections': defaultdict(int),
        'texts_with_pii': 0,
    }

    start = time.time()

    for csv_file in csv_files:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                text = row.get('Text', '')
                if not text:
                    continue

                total_stats['texts'] += 1
                detections = detect_pii(text)

                if detections:
                    total_stats['texts_with_pii'] += 1
                    for pii_type, matches in detections.items():
                        total_stats['detections'][pii_type] += len(matches)

    elapsed = time.time() - start

    return {
        'source': f'{len(csv_files)} generated files',
        'texts': total_stats['texts'],
        'texts_with_pii': total_stats['texts_with_pii'],
        'detections': dict(total_stats['detections']),
        'time': elapsed,
    }


def print_accuracy_results(result):
    """Print accuracy results in a formatted table."""
    print('='*65)
    print(f"RECALL BY PII TYPE - {result['source']}")
    print('='*65)
    print(f"\n{'PII Type':<15} {'Recall':<12} {'Detected':<10} {'Missed':<10} {'Total'}")
    print('-'*60)

    total_tp = 0
    total_fn = 0
    total_gt = 0

    accuracy = result['accuracy']
    for pii_type in ['EMAIL', 'PHONE', 'SSN', 'URL', 'CREDIT_CARD', 'PERSON', 'ADDRESS', 'COMPANY']:
        if pii_type not in accuracy:
            continue
        stats = accuracy[pii_type]
        tp, fn, gt = stats['tp'], stats['fn'], stats['total']
        total_tp += tp
        total_fn += fn
        total_gt += gt

        if gt > 0:
            recall = tp / gt * 100
            status = '✓' if recall >= 80 else '⚠️' if recall >= 50 else '✗'
            print(f'{pii_type:<15} {recall:>6.1f}% {status}    {tp:<10} {fn:<10} {gt}')
        else:
            print(f'{pii_type:<15} {"N/A":<12} {tp:<10} {fn:<10} {gt}')

    print('-'*60)
    overall = total_tp / total_gt * 100 if total_gt > 0 else 0
    print(f'{"OVERALL":<15} {overall:>6.1f}%      {total_tp:<10} {total_fn:<10} {total_gt}')
    print(f'\n📊 OVERALL RECALL: {overall:.1f}% ({total_tp}/{total_gt} PII items detected)')
    print(f"⏱️  Time: {result['time']:.2f}s | Samples: {result['samples']}")


def print_generated_results(result):
    """Print generated files benchmark results."""
    print('='*65)
    print(f"DETECTION STATS - {result['source']}")
    print('='*65)

    total = sum(result['detections'].values())
    rate = result['texts_with_pii'] / result['texts'] * 100 if result['texts'] > 0 else 0

    print(f"\nTexts processed: {result['texts']}")
    print(f"Texts with PII:  {result['texts_with_pii']} ({rate:.1f}%)")
    print(f"Total detections: {total}")
    print(f"\nBy type:")

    for pii_type, count in sorted(result['detections'].items(), key=lambda x: -x[1]):
        pct = count / total * 100 if total > 0 else 0
        print(f"  {pii_type:<15} {count:>6} ({pct:>5.1f}%)")

    print(f"\n⏱️  Time: {result['time']:.2f}s")


def read_pdf_text(pdf_path):
    """Extract text from PDF using pdftotext."""
    try:
        result = subprocess.run(
            ['pdftotext', '-layout', str(pdf_path), '-'],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            return result.stdout
    except FileNotFoundError:
        print("Warning: pdftotext not found. Install poppler-utils.")
    except subprocess.TimeoutExpired:
        print("Warning: PDF extraction timed out.")
    except Exception as e:
        print(f"Warning: PDF extraction failed: {e}")
    return None


def run_csv_pdf_comparison():
    """Compare PII detection between CSV and PDF versions of Testing_Set."""
    data_dir = Path(__file__).parent / 'data'
    csv_path = data_dir / 'Testing_Set.csv'
    pdf_path = data_dir / 'Testing_Set.pdf'

    results = {'csv': None, 'pdf': None}

    # Benchmark CSV
    if csv_path.exists():
        print("\nProcessing Testing_Set.csv...")
        start = time.time()

        texts = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('Text'):
                    texts.append(row['Text'])

        csv_detections = defaultdict(int)
        csv_samples = defaultdict(list)
        texts_with_pii = 0
        total_chars = sum(len(t) for t in texts)

        for text in texts:
            detections = detect_pii(text)
            if detections:
                texts_with_pii += 1
                for pii_type, matches in detections.items():
                    csv_detections[pii_type] += len(matches)
                    if len(csv_samples[pii_type]) < 2:
                        csv_samples[pii_type].extend(matches[:1])

        elapsed = time.time() - start
        results['csv'] = {
            'file': 'Testing_Set.csv',
            'texts': len(texts),
            'chars': total_chars,
            'texts_with_pii': texts_with_pii,
            'detections': dict(csv_detections),
            'total': sum(csv_detections.values()),
            'samples': dict(csv_samples),
            'time': elapsed,
        }

    # Benchmark PDF
    if pdf_path.exists():
        print("Processing Testing_Set.pdf...")
        start = time.time()

        pdf_text = read_pdf_text(pdf_path)
        if pdf_text:
            pdf_detections = defaultdict(int)
            pdf_samples = defaultdict(list)

            detections = detect_pii(pdf_text)
            for pii_type, matches in detections.items():
                pdf_detections[pii_type] += len(matches)
                if len(pdf_samples[pii_type]) < 2:
                    pdf_samples[pii_type].extend(matches[:1])

            elapsed = time.time() - start
            results['pdf'] = {
                'file': 'Testing_Set.pdf',
                'texts': 1,  # PDF is one continuous text
                'chars': len(pdf_text),
                'texts_with_pii': 1 if detections else 0,
                'detections': dict(pdf_detections),
                'total': sum(pdf_detections.values()),
                'samples': dict(pdf_samples),
                'time': elapsed,
            }

    return results


def print_csv_pdf_comparison(results):
    """Print CSV vs PDF comparison results."""
    print('='*75)
    print('CSV vs PDF PERFORMANCE COMPARISON')
    print('='*75)

    csv_result = results.get('csv')
    pdf_result = results.get('pdf')

    if not csv_result and not pdf_result:
        print("No results to compare.")
        return

    # Header
    print(f"\n{'Metric':<20} {'CSV':<25} {'PDF':<25}")
    print('-'*70)

    # File info
    if csv_result:
        print(f"{'File':<20} {csv_result['file']:<25}", end='')
    else:
        print(f"{'File':<20} {'N/A':<25}", end='')
    if pdf_result:
        print(f" {pdf_result['file']:<25}")
    else:
        print(f" {'N/A':<25}")

    # Text blocks
    csv_texts = csv_result['texts'] if csv_result else 0
    pdf_texts = pdf_result['texts'] if pdf_result else 0
    print(f"{'Text blocks':<20} {csv_texts:<25} {pdf_texts:<25}")

    # Characters
    csv_chars = f"{csv_result['chars']:,}" if csv_result else 'N/A'
    pdf_chars = f"{pdf_result['chars']:,}" if pdf_result else 'N/A'
    print(f"{'Characters':<20} {csv_chars:<25} {pdf_chars:<25}")

    # Total detections
    csv_total = csv_result['total'] if csv_result else 0
    pdf_total = pdf_result['total'] if pdf_result else 0
    print(f"{'Total detections':<20} {csv_total:<25} {pdf_total:<25}")

    # Time
    csv_time = f"{csv_result['time']:.3f}s" if csv_result else 'N/A'
    pdf_time = f"{pdf_result['time']:.3f}s" if pdf_result else 'N/A'
    print(f"{'Processing time':<20} {csv_time:<25} {pdf_time:<25}")

    # Throughput
    if csv_result and csv_result['time'] > 0:
        csv_throughput = f"{csv_result['chars'] / csv_result['time']:,.0f} chars/s"
    else:
        csv_throughput = 'N/A'
    if pdf_result and pdf_result['time'] > 0:
        pdf_throughput = f"{pdf_result['chars'] / pdf_result['time']:,.0f} chars/s"
    else:
        pdf_throughput = 'N/A'
    print(f"{'Throughput':<20} {csv_throughput:<25} {pdf_throughput:<25}")

    # Per-type comparison
    print(f"\n{'PII Type':<15} {'CSV Count':<15} {'PDF Count':<15} {'Diff':<15}")
    print('-'*60)

    all_types = set()
    if csv_result:
        all_types.update(csv_result['detections'].keys())
    if pdf_result:
        all_types.update(pdf_result['detections'].keys())

    for pii_type in sorted(all_types):
        csv_count = csv_result['detections'].get(pii_type, 0) if csv_result else 0
        pdf_count = pdf_result['detections'].get(pii_type, 0) if pdf_result else 0
        diff = pdf_count - csv_count
        diff_str = f"+{diff}" if diff > 0 else str(diff)

        # Show sample if available
        sample = ""
        if csv_result and pii_type in csv_result.get('samples', {}):
            s = csv_result['samples'][pii_type]
            if s:
                sample = f" (e.g., {s[0][:20]}...)" if len(s[0]) > 20 else f" (e.g., {s[0]})"

        print(f"{pii_type:<15} {csv_count:<15} {pdf_count:<15} {diff_str:<15}")

    print('-'*60)
    diff_total = pdf_total - csv_total
    diff_str = f"+{diff_total}" if diff_total > 0 else str(diff_total)
    print(f"{'TOTAL':<15} {csv_total:<15} {pdf_total:<15} {diff_str:<15}")

    # Summary
    print(f"\n📊 CSV detected {csv_total} PII items in {csv_texts} text blocks")
    print(f"📊 PDF detected {pdf_total} PII items from extracted text")

    if csv_total > 0 and pdf_total > 0:
        ratio = pdf_total / csv_total * 100
        print(f"📊 PDF/CSV ratio: {ratio:.1f}% ({pdf_total}/{csv_total})")


def main():
    parser = argparse.ArgumentParser(description='PII Detection Accuracy Benchmark')
    parser.add_argument('--full', action='store_true',
                       help='Run full test (Training_Set + Generated files + CSV/PDF comparison)')
    parser.add_argument('--samples', type=int, default=500,
                       help='Number of samples for Training_Set (default: 500)')
    args = parser.parse_args()

    mode = 'FULL' if args.full else 'QUICK'
    print('='*75)
    print(f'PII DETECTION ACCURACY BENCHMARK ({mode} TEST)')
    print('='*75)

    # Always run Training_Set benchmark
    training_result = run_training_set_benchmark(args.samples)
    if training_result:
        print_accuracy_results(training_result)

    # Run additional benchmarks for full test
    if args.full:
        # CSV vs PDF comparison
        print('\n')
        csv_pdf_results = run_csv_pdf_comparison()
        if csv_pdf_results:
            print_csv_pdf_comparison(csv_pdf_results)

        # Generated files benchmark
        print('\n')
        generated_result = run_generated_benchmark()
        if generated_result:
            print_generated_results(generated_result)

    print('\n' + '='*75)

    # Return exit code based on overall recall
    if training_result:
        accuracy = training_result['accuracy']
        total_tp = sum(a['tp'] for a in accuracy.values())
        total_gt = sum(a['total'] for a in accuracy.values())
        overall = total_tp / total_gt if total_gt > 0 else 0

        if overall >= 0.5:
            print('✓ BENCHMARK PASSED (>50% overall recall)')
            return 0
        else:
            print('✗ BENCHMARK FAILED (<50% overall recall)')
            return 1

    return 1


if __name__ == '__main__':
    sys.exit(main())
